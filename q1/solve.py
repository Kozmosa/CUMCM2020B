"""CLI entry: solve level 1 & 2, verify, write Result.xlsx."""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import List

from openpyxl import load_workbook

from .data import LevelConfig, load_level
from .dp import SolveResult, solve
from .model import DayRecord, VillagePurchaseMode
from .verify import verify_trajectory

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_RESULT = ROOT / "docs" / "2020B" / "Result.xlsx"
OUT_RESULT = ROOT / "q1" / "Result.xlsx"


def print_traj(cfg: LevelConfig, res: SolveResult) -> None:
    print(
        f"\n=== {cfg.name} | mode={res.purchase_mode.value} | "
        f"value={res.best_value:.1f} | arrival_day={res.arrival_day} | "
        f"end_cash={res.best_cash_at_end} W={res.best_water} F={res.best_food} ==="
    )
    print(f"{'day':>3} {'reg':>4} {'cash':>6} {'W':>4} {'F':>4}  action  weather buy")
    for r in res.trajectory:
        print(
            f"{r.day:3d} {r.region:4d} {r.cash:6d} {r.water:4d} {r.food:4d}  "
            f"{r.action:12s} {r.weather:9s} +W{r.bought_water}/F{r.bought_food}"
        )


def write_result_xlsx(
    path: Path,
    level1: List[DayRecord] | None,
    level2: List[DayRecord] | None,
    template: Path | None = None,
) -> None:
    """Fill the official Result.xlsx layout (cols A-E = L1, G-K = L2)."""
    if template and template.exists():
        wb = load_workbook(template)
        ws = wb.active
    else:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        headers = ("日期", "所在区域", "剩余资金数", "剩余水量", "剩余食物量")
        for i, h in enumerate(headers):
            ws.cell(3, 1 + i, h)
            ws.cell(3, 7 + i, h)

    def fill(traj: List[DayRecord], col_day: int) -> None:
        by_day = {r.day: r for r in traj}
        for day in range(0, 31):
            row = 4 + day  # day 0 -> row 4
            ws.cell(row, col_day, day)
            if day in by_day:
                r = by_day[day]
                ws.cell(row, col_day + 1, r.region)
                ws.cell(row, col_day + 2, r.cash)
                ws.cell(row, col_day + 3, r.water)
                ws.cell(row, col_day + 4, r.food)
            else:
                for k in range(1, 5):
                    ws.cell(row, col_day + k, None)

    if level1 is not None:
        fill(level1, 1)
    if level2 is not None:
        fill(level2, 7)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"Wrote {path}")


def run_level(which: str, mode: VillagePurchaseMode, verbose: bool) -> SolveResult:
    cfg = load_level(which)
    res = solve(cfg, purchase_mode=mode, verbose=verbose)
    print_traj(cfg, res)
    ok, msg = verify_trajectory(
        cfg, res.trajectory, mode, expected_value=res.best_value
    )
    print(f"verify: {msg}")
    if not ok:
        print("WARNING: verification failed", file=sys.stderr)
    return res


def main(argv: List[str] | None = None) -> int:
    p = argparse.ArgumentParser(description="Solve CUMCM 2020B Q1")
    p.add_argument("--level", default="both", choices=["1", "2", "both"])
    p.add_argument(
        "--purchase-mode",
        default=VillagePurchaseMode.START_OF_DAY.value,
        choices=[m.value for m in VillagePurchaseMode],
        help=(
            "village buy timing: "
            "start_of_day = buy only if already at village (Q1.md); "
            "after_arrival = buy after acting if ending at village"
        ),
    )
    p.add_argument("--quiet", action="store_true")
    p.add_argument("--out", type=Path, default=OUT_RESULT)
    p.add_argument("--template", type=Path, default=DEFAULT_RESULT)
    args = p.parse_args(argv)

    mode = VillagePurchaseMode(args.purchase_mode)
    verbose = not args.quiet

    r1 = r2 = None
    if args.level in ("1", "both"):
        r1 = run_level("1", mode, verbose)
    if args.level in ("2", "both"):
        r2 = run_level("2", mode, verbose)

    write_result_xlsx(
        args.out,
        r1.trajectory if r1 else None,
        r2.trajectory if r2 else None,
        template=args.template,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
