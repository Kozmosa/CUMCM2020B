"""CLI entry: solve Q2 levels 3 & 4, verify, print decision tables, run
sensitivity sweeps (failure penalty M and sandstorm probability), and write
Result.xlsx with the demonstration trajectories.

Usage:
    uv run python -m q2.solve                      # base run, both levels
    uv run python -m q2.solve --sensitivity        # + M / p_storm sweeps
    uv run python -m q2.solve --level 4 --quiet
"""

from __future__ import annotations

import argparse
import gc
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook

from .data import LevelConfig, Weather, level3, level4, load_level
from .dp import DEFAULT_M, SolveResult, solve
from .model import DayRecord, VillagePurchaseMode
from .policy import best_action
from .simulate import simulate
from .verify import verify_trajectory

ROOT = Path(__file__).resolve().parents[1]
OUT_RESULT = ROOT / "q2" / "Result.xlsx"

M_SWEEP = (1.0e3, 1.0e4, 1.0e6, 1.0e9)
P_STORM_SWEEP = (0.0, 0.05, 0.10, 0.15, 0.20)


# ---------------------------------------------------------------------------
# Reporting helpers
# ---------------------------------------------------------------------------


def print_traj(cfg: LevelConfig, res: SolveResult, traj: List[DayRecord], val: Optional[float]) -> None:
    print(
        f"\n=== {cfg.name} | mode={res.purchase_mode.value} | M={res.M:.0f} | "
        f"V0={res.V0:.1f} | P_succ={res.prob_succ:.5f} | q0={res.best_q0} | "
        f"realised={val} ==="
    )
    print(f"{'day':>3} {'reg':>4} {'cash':>6} {'W':>4} {'F':>4}  action  weather buy")
    for r in traj:
        print(
            f"{r.day:3d} {r.region:4d} {r.cash:6d} {r.water:4d} {r.food:4d}  "
            f"{r.action:12s} {r.weather:9s} +W{r.bought_water}/F{r.bought_food}"
        )


# Representative states for the decision table: (day, node, W, F).
def _probes(cfg: LevelConfig, q0: Tuple[int, int]) -> List[Tuple[int, int, int, int]]:
    if cfg.name == "level3":
        return [
            (1, 1, q0[0], q0[1]),   # first move out of the start
            (3, 9, 60, 60),         # at the mine, early
            (6, 9, 40, 40),         # at the mine, mid game
            (8, 9, 25, 25),         # at the mine, late / low supply
            (9, 12, 20, 20),        # next to the end on the last day
        ]
    return [
        (1, 1, q0[0], q0[1]),       # first move out of the start
        (6, 18, 150, 160),          # at the mine, plenty of supply
        (12, 18, 40, 45),           # at the mine, low supply
        (13, 14, 60, 60),           # at the village (buy decision)
        (13, 14, 200, 220),         # at the village, already stocked
        (21, 18, 100, 100),         # at the mine, late game
        (27, 20, 40, 40),           # near the end, late game
    ]


def print_decision_table(cfg: LevelConfig, res: SolveResult) -> None:
    """Optimal action at representative states, one row per (state, weather)."""
    thetas = [k for k, p in cfg.p_weather.items() if p > 0]
    print(f"\n--- decision table ({cfg.name}) ---")
    print(f"{'day':>3} {'node':>4} {'W':>4} {'F':>4}  weather   action (+buy)")
    for day, v, w, f in _probes(cfg, res.best_q0):
        for th in thetas:
            ch = best_action(cfg, res, day, v, w, f, th)
            buy = f"+W{ch.buy_w}/F{ch.buy_f}" if (ch.buy_w or ch.buy_f) else ""
            print(f"{day:3d} {v:4d} {w:4d} {f:4d}  {th:9s} {ch.action}{buy}")


def write_result_xlsx(
    path: Path,
    level3_traj: List[DayRecord] | None,
    level4_traj: List[DayRecord] | None,
) -> None:
    """Write the official-style Result.xlsx (cols A-E = L3, G-K = L4)."""
    wb = Workbook()
    ws = wb.active
    ws.cell(1, 3, "第三关")
    ws.cell(1, 9, "第四关")
    headers = ("日期", "所在区域", "剩余资金数", "剩余水量", "剩余食物量")
    for i, h in enumerate(headers):
        ws.cell(3, 1 + i, h)
        ws.cell(3, 7 + i, h)

    def fill(traj: List[DayRecord], col_day: int) -> None:
        by_day = {r.day: r for r in traj}
        for day in range(0, 31):
            row = 4 + day  # day 0 -> row 4
            ws.cell(row, col_day, day)
            if day in by_day:
                r = by_day[day]
                ws.cell(row, col_day + 1, r.region)
                ws.cell(row, col_day + 2, r.cash)
                ws.cell(row, col_day + 3, r.water)
                ws.cell(row, col_day + 4, r.food)

    if level3_traj is not None:
        fill(level3_traj, 1)
    if level4_traj is not None:
        fill(level4_traj, 7)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"Wrote {path}")


# ---------------------------------------------------------------------------
# Runs
# ---------------------------------------------------------------------------


def run_level(
    cfg: LevelConfig,
    mode: VillagePurchaseMode,
    M: float,
    verbose: bool,
    show_table: bool = True,
) -> Tuple[SolveResult, List[DayRecord], Optional[float]]:
    res = solve(cfg, purchase_mode=mode, M=M, verbose=verbose)
    traj, val = simulate(cfg, res)
    ok, msg = verify_trajectory(cfg, traj, mode, res=res)
    print_traj(cfg, res, traj, val)
    print(f"verify: {msg}")
    if not ok:
        print("WARNING: verification failed", file=sys.stderr)
    if show_table:
        print_decision_table(cfg, res)
    return res, traj, val


def sweep_M(cfg: LevelConfig, mode: VillagePurchaseMode) -> None:
    print(f"\n--- sensitivity: failure penalty M ({cfg.name}) ---")
    print(f"{'M':>10} {'V0':>12} {'P_succ':>9}  q0")
    for M in M_SWEEP:
        res = solve(cfg, purchase_mode=mode, M=M, verbose=False)
        print(f"{M:10.0f} {res.V0:12.1f} {res.prob_succ:9.5f}  {res.best_q0}")
        del res
        gc.collect()


def sweep_p_storm(mode: VillagePurchaseMode, M: float) -> None:
    """L4: vary p_storm, keeping sunny:hot = 5:4 among the rest."""
    print(f"\n--- sensitivity: p_storm (level4, sunny:hot = 5:4) ---")
    print(f"{'p_storm':>7} {'V0':>12} {'P_succ':>9}  q0")
    for ps in P_STORM_SWEEP:
        p = {
            "sunny": (1 - ps) * 5 / 9,
            "hot": (1 - ps) * 4 / 9,
            "sandstorm": ps,
        }
        cfg = level4(p_weather=p)
        res = solve(cfg, purchase_mode=mode, M=M, verbose=False)
        print(f"{ps:7.2f} {res.V0:12.1f} {res.prob_succ:9.5f}  {res.best_q0}")
        del res, cfg
        gc.collect()


def main(argv: List[str] | None = None) -> int:
    p = argparse.ArgumentParser(description="Solve CUMCM 2020B Q2")
    p.add_argument("--level", default="both", choices=["3", "4", "both"])
    p.add_argument(
        "--purchase-mode",
        default=VillagePurchaseMode.START_OF_DAY.value,
        choices=[m.value for m in VillagePurchaseMode],
        help=(
            "village buy timing: "
            "start_of_day = buy if starting the day at a village (Q1 semantics); "
            "after_arrival = buy after acting if ending the day at a village"
        ),
    )
    p.add_argument("--M", type=float, default=DEFAULT_M, help="failure penalty")
    p.add_argument("--quiet", action="store_true")
    p.add_argument("--no-sim", action="store_true", help="skip trajectory/verify")
    p.add_argument("--no-table", action="store_true", help="skip decision table")
    p.add_argument("--sensitivity", action="store_true", help="run M / p_storm sweeps")
    p.add_argument("--out", type=Path, default=OUT_RESULT)
    args = p.parse_args(argv)

    mode = VillagePurchaseMode(args.purchase_mode)
    verbose = not args.quiet

    trajs: Dict[str, List[DayRecord]] = {}
    for which in (("3", "4") if args.level == "both" else (args.level,)):
        cfg = load_level(which)
        if args.no_sim:
            res = solve(cfg, purchase_mode=mode, M=args.M, verbose=verbose)
            print(
                f"{cfg.name}: V0={res.V0:.1f} P_succ={res.prob_succ:.5f} q0={res.best_q0}"
            )
        else:
            _, traj, _ = run_level(cfg, mode, args.M, verbose, not args.no_table)
            trajs[cfg.name] = traj

    if not args.no_sim and trajs:
        write_result_xlsx(args.out, trajs.get("level3"), trajs.get("level4"))

    if args.sensitivity:
        for which in (("3", "4") if args.level == "both" else (args.level,)):
            sweep_M(load_level(which), mode)
        if args.level in ("4", "both"):
            sweep_p_storm(mode, args.M)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
